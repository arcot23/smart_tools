import csv
import os
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
from collections import defaultdict
import argparse

def summarize_xml_file(xml_path):
    """
    Summarizes an XML file by listing unique tag paths and their occurrence counts.

    Args:
        xml_path (str): Full path to the XML file.

    Returns:
        dict: A dictionary with tag paths as keys and occurrence counts as values.
    """
    def traverse(element, path, counter):
        tag_path = f"{path}\\{element.tag}" if path else element.tag
        counter[tag_path] += 1

        # Count attributes
        for attr in element.attrib:
            attr_path = f"{tag_path}\\@{attr}"
            counter[attr_path] += 1

        # Recurse into children
        for child in element:
            traverse(child, tag_path, counter)

    counter = defaultdict(int)

    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()
        traverse(root, "", counter)
    except Exception as e:
        print(f"Error processing XML file: {e}")
        return {}

    return dict(counter)

def summarize_excel_file(excel_path, encoding=None):
    """
    Summarizes all sheets in an Excel (.xlsx) file.

    Args:
        excel_path (str): Full path to the Excel file.
        encoding (str): Ignored for Excel files, kept for interface consistency.

    Returns:
        list: A list of dictionaries, each summarizing one sheet.
    """
    summaries = []

    try:
        wb = load_workbook(excel_path, read_only=True)
        for sheet in wb.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue

            header = rows[0]
            sheet_summary = {
                "sheet_name": sheet.title,
                "row_count": len(rows) - 1,  # exclude header
                "column_count": len(header),
                "column_names": list(header)
            }
            summaries.append(sheet_summary)

    except Exception as e:
        print(f"Error processing Excel file: {e}")

    return summaries

def summarize_csv_file(csv_path, encoding):
    """
    Summarizes a CSV file by returning row count, column count, and column names.

    Args:
        csv_path (str): Full path to the CSV file.
        encoding (str): Encoding used to read the file.

    Returns:
        dict: Summary with keys 'row_count', 'column_count', and 'column_names'.
    """
    summary = {
        "delimiter": "",
        "row_count": 0,
        "column_count": 0,
        "column_names": [],
    }

    delimiters = [",", "\t", ";", "|"]

    try:
        for index, delimiter in enumerate(delimiters):
            with open(csv_path, mode="r", encoding=encoding, newline='') as f:
                reader = csv.reader(f, delimiter = delimiter)
                header = next(reader, None)
                if header:
                    if len(header) == 1 and index < len(delimiters):
                        continue
                    summary["delimiter"] = delimiter
                    summary["column_names"] = header
                    summary["column_count"] = len(header)
                summary["row_count"] = sum(1 for _ in reader)
            return summary
    except Exception as e:
        print(f"Error processing CSV: {e}")

    return summary


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Summarize structured data files.")
    parser.add_argument("file", help="Path to the input file")
    parser.add_argument(
        "--type",
        choices=["csv", "excel", "xml"],
        required=True,
        help="Type of the input file"
    )
    parser.add_argument(
        "--encoding",
        default="utf-8",
        help="Encoding for CSV files (default: utf-8)"
    )

    args = parser.parse_args()
    file_path = args.file
    content_type = args.type

    if not os.path.exists(file_path):
        print(f"Error: File not found at {file_path}")
        exit(1)

    if content_type == "csv":
        print(summarize_csv_file(file_path, args.encoding))
    elif content_type == "excel":
        print(summarize_excel_file(file_path))
    elif content_type == "xml":
        print(summarize_xml_file(file_path))
